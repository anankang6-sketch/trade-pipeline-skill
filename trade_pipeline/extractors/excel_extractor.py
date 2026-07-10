"""
extractors/excel_extractor.py — Excel 询价单提取器

支持两种格式：
  - standard: 标准紧固件 (COMMODITY DESCRIPTION, Barcode, PCS...)
  - washers_mar: 垫圈 (Size, Cartons, Pallets, Qty tons...)

输出 ExtractedDocument，接入 L1 ExtractorCache。
"""
import re
from dataclasses import asdict
from pathlib import Path
from openpyxl import load_workbook

from ..extractors.base import ExtractedDocument, ContentBlock
from ..understanding.cache import ExtractorCache

EXTRACTOR_VERSION = "1.0"

# 读取上限（T6）：防御异常大 / 恶意膨胀的 Excel 拖垮内存或耗尽 CPU
# （zip bomb、超大 sheet）。超限直接抛清晰异常，让调用方给出可读提示。
MAX_FILE_SIZE_BYTES = 20 * 1024 * 1024   # 20 MB
MAX_ROWS = 5000


class FileTooLargeError(Exception):
    """询价单文件超过大小 / 行数上限。"""


def _check_file_size(path: Path) -> None:
    """校验磁盘文件大小不超过 MAX_FILE_SIZE_BYTES。"""
    size = path.stat().st_size
    if size > MAX_FILE_SIZE_BYTES:
        raise FileTooLargeError(
            f"询价单文件过大：{size / 1024 / 1024:.1f} MB，"
            f"超过上限 {MAX_FILE_SIZE_BYTES // 1024 // 1024} MB。"
            "请确认文件是否为正常询价单（可拆分或另存为精简 xlsx 后重试）。"
        )


def _check_row_count(ws) -> None:
    """校验 sheet 行数不超过 MAX_ROWS。"""
    if ws.max_row > MAX_ROWS:
        raise FileTooLargeError(
            f"询价单行数过多：{ws.max_row} 行，超过上限 {MAX_ROWS} 行。"
            "请确认是否包含大量空行 / 无关数据（清理后重试，或拆分为多个订单）。"
        )


def detect_format(ws) -> str:
    """检测询价单格式: 'standard' 或 'washers_mar'"""
    parts = []
    for r in range(1, min(ws.max_row + 1, 6)):
        for c in range(1, min(ws.max_column + 1, 25)):
            val = ws.cell(r, c).value
            if val:
                parts.append(str(val))
    all_text = " ".join(parts)

    if "COMMODITY DESCRIPTION" in all_text or "Barcode" in all_text:
        return "standard"
    if "Size" in all_text and ("tons" in all_text.lower() or "Qty" in all_text):
        return "washers_mar"
    return "standard"


def _score_sheet(ws) -> int:
    """Score a sheet on how likely it is an inquiry/RFQ."""
    score = 0
    name = ws.title

    if "询价" in name or "RFQ" in name.upper() or "INQUIRY" in name.upper():
        score += 10
    if re.match(r'^Sheet\d+$', name, re.IGNORECASE):
        score -= 2

    for r in range(1, min(ws.max_row + 1, 6)):
        text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column + 1, 25)))
        if "询价" in text or "INQUIRY" in text.upper() or "RFQ" in text.upper():
            score += 8
        if "报价" in text and ("单价" in text or "PRICE" in text.upper()):
            score += 5
        if "COMMODITY DESCRIPTION" in text.upper():
            score += 5
    return score


def select_best_sheet(wb):
    """Select the best inquiry sheet from workbook."""
    if len(wb.sheetnames) == 1:
        return wb.active

    best_ws = wb.active
    best_score = _score_sheet(best_ws)

    for name in wb.sheetnames:
        ws = wb[name]
        s = _score_sheet(ws)
        if s > best_score:
            best_score = s
            best_ws = ws

    return best_ws


def _extract_supplementary_prices(wb, main_ws) -> dict[str, float]:
    """扫描非主 Sheet，按 barcode 提取价格数据"""
    prices = {}
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.title == main_ws.title:
            continue
        # 扫描前 5 行找含 "price" 或 "价格" 的列头
        price_col = None
        barcode_col = None
        header_row = None
        for r in range(1, min(ws.max_row + 1, 6)):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(r, c).value or "").lower().replace("\n", " ")
                if "产品编码" in val or "barcode" in val:
                    barcode_col = c
                    header_row = r
                if ("价格" in val or "price" in val) and "unit" not in val:
                    price_col = c
                    header_row = r
            if price_col and barcode_col:
                break
        if not price_col or not barcode_col or not header_row:
            continue
        # 读取数据行
        for r in range(header_row + 1, ws.max_row + 1):
            bc = ws.cell(r, barcode_col).value
            pv = ws.cell(r, price_col).value
            if bc and pv:
                bc_str = str(bc).strip()
                try:
                    prices[bc_str] = float(pv)
                except (ValueError, TypeError):
                    pass
    return prices


def extract(file_path: str, cache_dir: str | None = None) -> ExtractedDocument:
    """
    提取 Excel 询价单为 ExtractedDocument。

    参数:
        file_path: Excel 文件路径
        cache_dir: L1 缓存目录（None 则不缓存）

    返回:
        ExtractedDocument
    """
    path = Path(file_path)

    # 读取上限：先查文件大小（T6），避免把超大文件读进 openpyxl
    _check_file_size(path)

    # L1 缓存检查
    cache = None
    file_hash = None
    if cache_dir:
        cache = ExtractorCache(cache_dir, EXTRACTOR_VERSION)
        file_hash = cache.compute_file_hash(str(path))
        cached = cache.get(file_hash)
        if cached:
            # 从缓存重建
            blocks = [ContentBlock(**b) for b in cached.get("blocks", [])]
            return ExtractedDocument(
                content_text=cached["content_text"],
                blocks=blocks,
                meta=cached["meta"],
                confidence=cached["confidence"],
                source_path=cached["source_path"],
                extraction_method=cached["extraction_method"],
                warnings=cached.get("warnings", []),
                content_hash=cached.get("content_hash", ""),
            )

    wb = load_workbook(str(path), data_only=True)
    ws = select_best_sheet(wb)
    # 读取上限：选定 sheet 后查行数（T6）。read_only 与既有代码不兼容
    # （detect_format / _score_sheet / iter_rows 会多次遍历并随机访问 cell，
    # read_only 模式下这些会失效或性能反劣），故只加上限检查，不切 read_only。
    _check_row_count(ws)
    fmt = detect_format(ws)

    # 构建文本和表格数据
    lines = []
    table_rows = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [str(c.value) if c.value is not None else "" for c in row]
        lines.append("\t".join(vals))
        table_rows.append(vals)

    content_text = "\n".join(lines)

    blocks = [
        ContentBlock(
            block_type="table",
            content=content_text,
            rows=table_rows,
            lang_hint="en",
        )
    ]

    doc = ExtractedDocument(
        content_text=content_text,
        blocks=blocks,
        meta={
            "filename": path.name,
            "format": fmt,
            "sheet": ws.title,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "supplementary_prices": _extract_supplementary_prices(wb, ws),
        },
        confidence=1.0,
        source_path=str(path),
        extraction_method="openpyxl",
    )

    wb.close()

    # L1 缓存写入
    if cache and file_hash:
        doc_dict = asdict(doc)
        cache.put(file_hash, doc_dict)

    return doc
