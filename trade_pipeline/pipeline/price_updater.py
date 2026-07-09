"""
pipeline/price_updater.py — 报价单单价回写（item_uuid 锚点定位）

禁止使用行号/header offset 方式回写。
通过扫描隐藏列的 item_uuid 定位每行，再读取对应单价列。
"""
from openpyxl import load_workbook

UUID_COL_HEADER = "__item_uuid__"


class PriceUpdateError(Exception):
    """价格回写过程中的致命错误"""
    pass


def _find_uuid_column(ws) -> int | None:
    """扫描第一行~第 20 行，找到含 UUID_COL_HEADER 的列"""
    for row in range(1, min(ws.max_row + 1, 21)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val and str(val).strip() == UUID_COL_HEADER:
                return col
    return None


def _find_price_column(ws, uuid_col: int) -> int | None:
    """找到单价列（列头含 'Unit Price' 或 'Price'）。找不到返回 None。

    修复（遗漏4）：原实现找不到时 fallback 到硬编码第 6 列(F)，会在用户
    调整了报价单列顺序时静默从错误列读数当单价回写——这正是本模块开头
    禁止的"行号/offset 方式"。改为返回 None，由调用方硬阻断。
    """
    for row in range(1, min(ws.max_row + 1, 21)):
        if ws.cell(row, uuid_col).value == UUID_COL_HEADER:
            # 同一行扫描其他列找 Price
            for col in range(1, uuid_col):
                val = str(ws.cell(row, col).value or "")
                if "price" in val.lower() and "unit" in val.lower():
                    return col
                if "price" in val.lower():
                    return col
            break
    return None


def update_prices(model, quotation_path: str) -> dict:
    """
    从已填单价的报价单 Excel 读取单价，通过 item_uuid 锚点定位回写到 OrderModel。

    参数:
        model: OrderModel 实例
        quotation_path: 报价单 Excel 路径

    返回:
        {"updated": N, "errors": [...], "warnings": [...]}

    保证:
        即使报价单中插入分组标题行、调整顺序、插入空行，也不会导致单价错位回写。
    """
    wb = load_workbook(quotation_path, data_only=True)
    ws = wb.active

    # 找 uuid 列
    uuid_col = _find_uuid_column(ws)
    if uuid_col is None:
        raise PriceUpdateError(
            f"报价单 '{quotation_path}' 中未找到 uuid 隐藏列 ('{UUID_COL_HEADER}')。"
            f"请确认该报价单由 quote_writer 生成。"
        )

    # 找单价列
    price_col = _find_price_column(ws, uuid_col)
    if price_col is None:
        raise PriceUpdateError(
            f"报价单 '{quotation_path}' 中未找到单价列（列头需含 'Price'）。"
            f"请确认该报价单由 quote_writer 生成且单价列表头未被删改。"
        )

    # 构建 uuid → Excel 行号映射
    uuid_to_row: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row, uuid_col).value
        if cell_val and str(cell_val).strip() and str(cell_val).strip() != UUID_COL_HEADER:
            uid = str(cell_val).strip()
            if uid in uuid_to_row:
                raise PriceUpdateError(f"uuid 重复: '{uid}' 出现在行 {uuid_to_row[uid]} 和 {row}")
            uuid_to_row[uid] = row

    # 构建 uuid → model item 映射
    uuid_to_item: dict[str, object] = {}
    for item in model.items:
        if item.item_uuid in uuid_to_item:
            raise PriceUpdateError(f"模型中 uuid 重复: '{item.item_uuid}'")
        uuid_to_item[item.item_uuid] = item

    # 交叉校验
    errors = []
    warnings = []
    model_uuids = set(uuid_to_item.keys())
    excel_uuids = set(uuid_to_row.keys())

    orphan_model = model_uuids - excel_uuids
    orphan_excel = excel_uuids - model_uuids

    if orphan_model:
        errors.append(
            f"模型中 {len(orphan_model)} 个 item 在 Excel 中未找到: "
            f"{sorted(orphan_model)}"
        )
    if orphan_excel:
        warnings.append(
            f"Excel 中 {len(orphan_excel)} 个 uuid 不在模型中（可能已删除）: "
            f"{sorted(orphan_excel)}"
        )

    # 读取单价并回写到 model
    updated = 0
    for uid, row_num in uuid_to_row.items():
        if uid not in uuid_to_item:
            continue
        price = ws.cell(row_num, price_col).value
        if price is None or str(price).strip() == "":
            warnings.append(f"item {uid} (行{row_num}) 单价为空，跳过")
            continue
        try:
            price_val = float(price)
        except (ValueError, TypeError):
            errors.append(f"item {uid} (行{row_num}) 单价格式非法: '{price}'")
            continue
        if price_val < 0:
            errors.append(f"item {uid} (行{row_num}) 单价为负数: {price_val}")
            continue

        uuid_to_item[uid].unit_price = price_val
        updated += 1

    wb.close()
    return {"updated": updated, "errors": errors, "warnings": warnings}
