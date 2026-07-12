"""
writers/quote_writer.py — Quotation generator with UUID anchor column

Includes seller header, product table with hidden UUID column,
and footer terms section (matching PI/CI/PL layout).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

from .base_writer import BaseWriter, _sanitize

FONT_NAME = "Calibri"
HEADER_COLOR = "1F3864"
BORDER_COLOR = "E0E0E0"
UUID_COL_HEADER = "__item_uuid__"


def _fnt(size=10, bold=False, color="222222", name=FONT_NAME):
    return Font(name=name, size=size, bold=bold, color=color)


def _aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _thin(color=BORDER_COLOR):
    return Side(style="thin", color=color)


def _mc(ws, r, c1, c2, value=None, font=None, align=None):
    if c1 != c2:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1)
    if value is not None:
        # 公式注入净化（T3）：quote_writer 不走 base_writer.sc，需自行净化。
        cell.value = _sanitize(value, formula=False)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    return cell


class QuoteWriter(BaseWriter):
    """Quotation Writer with UUID anchor column for price write-back."""

    def write(self, output_path: str, **kwargs) -> dict:
        model = self.model
        items = model.items
        order_no = model.order.order_no
        currency = model.order.currency or "CNY"
        price_unit = model.order.price_unit or f"{currency}/SQM"
        date = model.order.date
        has_weight = model.derived.has_weight or False

        seller = self.seller
        buyer = self.buyer
        terms = self.terms

        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"

        NCOLS = 8
        LAST_COL_IDX = 8
        uuid_col_idx = NCOLS + 1

        if has_weight:
            HEADERS = ["No.", "Barcode", "Description", "UOM", "Qty (m²)",
                       f"Unit Price\n({price_unit})", "Weight\n(KG)", "Qty/Roll"]
            col_widths = [5, 16, 50, 5.5, 12, 16, 13, 8]
        else:
            HEADERS = ["No.", "Barcode", "Description", "UOM", "Qty (m²)",
                       f"Unit Price\n({price_unit})", f"Amount\n({currency})", "Qty/Roll"]
            col_widths = [5, 16, 50, 5.5, 12, 16, 16, 8]

        ws.sheet_properties = WorksheetProperties(
            pageSetUpPr=PageSetupProperties(fitToPage=True))
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        uuid_letter = get_column_letter(uuid_col_idx)
        ws.column_dimensions[uuid_letter].width = 0
        ws.column_dimensions[uuid_letter].hidden = True

        row = 1

        # ── Seller header ──────────────────────────────────────
        seller_name_cn = seller.get("name_cn", "")
        seller_name_en = seller.get("name_en", "")
        seller_address = seller.get("address", "")
        seller_contact = seller.get("contact", "")
        seller_email = seller.get("email", "")
        seller_tel = seller.get("tel", "")

        if seller_name_cn:
            ws.row_dimensions[row].height = 24
            _mc(ws, row, 1, LAST_COL_IDX, value=seller_name_cn,
                font=_fnt(18, True, HEADER_COLOR, "宋体"), align=_aln("center"))
            row += 1

        if seller_name_en:
            ws.row_dimensions[row].height = 20
            _mc(ws, row, 1, LAST_COL_IDX, value=seller_name_en,
                font=_fnt(14, True, HEADER_COLOR), align=_aln("center"))
            row += 1

        if seller_address:
            ws.row_dimensions[row].height = 14
            _mc(ws, row, 1, LAST_COL_IDX, value=seller_address,
                font=_fnt(9, color="666666"), align=_aln("center"))
            row += 1

        # Separator
        ws.row_dimensions[row].height = 4
        _mc(ws, row, 1, LAST_COL_IDX)
        ws.cell(row, 1).fill = _fill(HEADER_COLOR)
        row += 1

        # ── QUOTATION title ────────────────────────────────────
        ws.row_dimensions[row].height = 28
        _mc(ws, row, 1, LAST_COL_IDX, value="QUOTATION",
            font=_fnt(16, True, HEADER_COLOR), align=_aln("center", "center"))
        row += 1

        # Quote No. + Date
        ws.row_dimensions[row].height = 16
        ws.cell(row, 1, "Quote No.:").font = _fnt(9, True, HEADER_COLOR)
        ws.cell(row, 2, f"QT-{order_no}").font = _fnt(9)
        if date:
            ws.cell(row, 7, "Date:").font = _fnt(9, True, HEADER_COLOR)
            ws.cell(row, 7).alignment = _aln("right")
            ws.cell(row, 8, date).font = _fnt(9)
        row += 1

        # Contact info
        if seller_tel or seller_email:
            ws.row_dimensions[row].height = 14
            parts = []
            if seller_tel:
                parts.append(f"Tel: {seller_tel}")
            if seller_contact:
                parts.append(f"Contact: {seller_contact}")
            if seller_email:
                parts.append(f"Email: {seller_email}")
            ws.cell(row, 1, "  |  ".join(parts)).font = _fnt(8, color="666666")
            row += 1

        # Buyer
        buyer_name = buyer.get("name_en", "") or buyer.get("name_ru", "")
        if buyer_name:
            ws.row_dimensions[row].height = 16
            ws.cell(row, 1, "To:").font = _fnt(9, True)
            # 公式注入净化（T3）：买方名可能含 = / @ 等触发字符
            ws.cell(row, 2, _sanitize(buyer_name, formula=False)).font = _fnt(9, True)
            row += 1

        ws.row_dimensions[row].height = 6
        row += 1

        # ── Column headers ─────────────────────────────────────
        ws.row_dimensions[row].height = 32
        for col_i, h in enumerate(HEADERS, 1):
            c = ws.cell(row, col_i, h)
            c.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
            c.fill = _fill(HEADER_COLOR)
            c.alignment = _aln("center", "center", wrap=True)

        c = ws.cell(row, uuid_col_idx, UUID_COL_HEADER)
        c.font = Font(name=FONT_NAME, size=1, color="FFFFFF")
        header_row = row
        row += 1

        # ── Data rows ──────────────────────────────────────────
        data_start_row = row
        no = 1
        for item in items:
            ws.row_dimensions[row].height = 15
            vals = [no, getattr(item, "barcode", ""), item.description,
                    getattr(item, "unit", "pcs"), item.quantity,
                    getattr(item, "unit_price", None),
                    getattr(item, "weight_kg", None),
                    getattr(item, "qty_box", None)]

            for col_i, v in enumerate(vals, 1):
                # 公式注入净化（T3）：description/barcode 等来自输入数据的字符串
                c = ws.cell(row, col_i, _sanitize(v, formula=False))
                c.font = _fnt()
                c.border = Border(bottom=_thin())
                if col_i == 5:
                    c.number_format = "#,##0.##"
                elif col_i in (6, 7):
                    c.number_format = "#,##0.00"
                if col_i == 6:
                    c.fill = _fill("FFFBE6")
                    c.border = Border(
                        bottom=_thin(),
                        left=_thin("FFD700"),
                        right=_thin("FFD700"),
                    )

            uuid_cell = ws.cell(row, uuid_col_idx, item.item_uuid)
            uuid_cell.font = Font(name=FONT_NAME, size=1, color="FFFFFF")
            no += 1
            row += 1

        # ── Footer terms ───────────────────────────────────────
        if terms:
            row += 1
            ws.row_dimensions[row].height = 4
            _mc(ws, row, 1, LAST_COL_IDX)
            ws.cell(row, 1).fill = _fill(HEADER_COLOR)
            row += 1
            row += 1

            terms_items = [
                ("Payment Terms:", terms.get("payment", "")),
                ("Delivery:", terms.get("delivery", "")),
                ("Lead Time:", terms.get("lead_time", "")),
                ("Validity:", terms.get("validity", "")),
                ("Packing:", terms.get("packing", "")),
                ("Quality:", terms.get("quality", "")),
            ]
            if terms.get("exchange_rate"):
                terms_items.append(("Exchange Rate:", terms["exchange_rate"]))

            for lbl, val in terms_items:
                if not val:
                    continue
                ws.cell(row, 1, lbl).font = _fnt(9, True, HEADER_COLOR)
                ws.cell(row, 1).alignment = _aln("left", "top")
                _mc(ws, row, 3, LAST_COL_IDX, value=val,
                    font=_fnt(9), align=_aln("left", "top", wrap=True))
                ws.row_dimensions[row].height = 16
                row += 1

        wb.save(output_path)
        return {
            "items": len(items),
            "uuid_col": uuid_col_idx,
            "uuid_col_letter": uuid_letter,
            "header_row": header_row,
            "data_start_row": data_start_row,
        }


def write_quotation_with_uuid(items, out_path, order_no, has_weight=False,
                              currency="CNY", price_unit="CNY/MPCS",
                              seller=None, buyer=None, terms=None, date=None):
    """Legacy function interface — delegates to QuoteWriter."""
    from trade_pipeline.models.order_model import (
        OrderModel, OrderRefs, OrderInfo, DerivedData,
        ResolvedEntities, OrderMeta,
    )
    model = OrderModel(
        refs=OrderRefs(seller_id="", buyer_id="", terms_id=""),
        order=OrderInfo(
            order_no=order_no, format="", quote_no=f"QT-{order_no}",
            pi_number="", ci_number="", date=date or "",
            currency=currency, price_unit=price_unit,
        ),
        items=items,
        derived=DerivedData(
            total_items=len(items), total_qty=0, has_weight=has_weight,
        ),
        resolved=ResolvedEntities(
            seller=seller or {}, buyer=buyer or {}, terms=terms or {}, bank={},
        ),
        meta=OrderMeta(source_files=[]),
    )
    writer = QuoteWriter(model)
    return writer.write(out_path)
