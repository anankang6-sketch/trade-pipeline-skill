"""
writers/pi_writer.py — PI 形式发票生成器

从 OrderModel 读取数据，生成 PI Excel。
继承 BaseWriter 统一接口。

禁止：
  - 在 Writer 里重新解析 Excel
  - 绕过 OrderModel 直接读取
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .base_writer import BaseWriter, sc as _sc, mc as _mc, brd as _brd, rh as _rh
from ..models.amounts import amount_formula, PricingMode

CALIBRI = "Calibri"


def _fnt(name=CALIBRI, size=11, bold=False, color="000000"):
    return Font(name=name, size=size, bold=bold, color=color)


def _aln(h="left", v=None, wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


class PIWriter(BaseWriter):
    """PI 形式发票 Writer"""

    def write(self, output_path: str, **kwargs) -> dict:
        """
        生成 PI Excel。

        返回:
            {"items": N, "total_qty": N, "output_path": str}
        """
        model = self.model
        items = model.items

        wb = Workbook()
        ws = wb.active
        ws.title = "PI"

        # 列宽（15列 A:O）
        col_widths = [8, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 12, 12, 12, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 页面设置
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

        today = model.order.date
        R = 1

        # ── 表头区 ──────────────────────────────────────────
        seller_name_cn = self.seller.get("name_cn", "")
        seller_name_en = self.seller.get("name_en", "")
        seller_address = self.seller.get("address", "")
        seller_contact = self.seller.get("contact", "")
        seller_email = self.seller.get("email", "")
        seller_tel = self.seller.get("tel", "")

        buyer_name = self.buyer.get("name_en", "") or self.buyer.get("name_ru", "")
        buyer_address_lines = self.buyer.get("address_lines", [])
        buyer_address = self.buyer.get("address", "")

        # Row 1: 卖方中文名
        if seller_name_cn:
            _mc(ws, R, 1, R, 15, value=seller_name_cn,
                font=_fnt("宋体", 20, bold=True), align=_aln("center"))
            _rh(ws, R, 21)
            R += 1

        # Row 2: 卖方英文名
        _mc(ws, R, 1, R, 15, value=seller_name_en,
            font=_fnt(CALIBRI, 14, bold=True), align=_aln("center"))
        _rh(ws, R, 18)
        R += 1

        # Row 3: 地址
        if seller_address:
            _mc(ws, R, 1, R, 15, value=seller_address,
                font=_fnt(CALIBRI, 10), align=_aln("center"))
            _rh(ws, R, 26)
            R += 1

        # Row 4: 空行
        _rh(ws, R, 12)
        R += 1

        # Row 5: PROFORMA INVOICE
        _mc(ws, R, 1, R, 15, value="PROFORMA INVOICE",
            font=_fnt(CALIBRI, 14, bold=True), align=_aln("center"))
        _rh(ws, R, 21)
        R += 1

        # Row 6: 空行（保证 PI No. 落在 Row 7 = O7，与 pl-gen 兼容）
        _rh(ws, R, 6)
        R += 1

        # Row 7: 联系方式 + PI 编号（必须在 Row 7，pl-gen 从 O7 读取）
        _sc(ws, R, 1, value=f"Tel: {seller_tel}",
            font=_fnt(CALIBRI, 10), align=_aln("left"))
        _sc(ws, R, 14, value="PI No.", font=_fnt(CALIBRI, 10))
        _sc(ws, R, 15, value=model.order.pi_number,
            font=_fnt(CALIBRI, 10, bold=True))
        R += 1

        # Row 8: Contact
        _sc(ws, R, 1, value=f"Contact: {seller_contact}",
            font=_fnt(CALIBRI, 10), align=_aln("left"))
        R += 1

        # Row 9: Email + Date
        _sc(ws, R, 1, value=f"E-mail: {seller_email}",
            font=_fnt(CALIBRI, 10), align=_aln("left"))
        _sc(ws, R, 14, value="Date:", font=_fnt(CALIBRI, 10))
        _sc(ws, R, 15, value=today, font=_fnt(CALIBRI, 10))
        R += 1

        # Row 10: 空行
        _rh(ws, R, 8)
        R += 1

        # Row 11: To (buyer)
        _sc(ws, R, 1, value="To:", font=_fnt(CALIBRI, 10, bold=True))
        _sc(ws, R, 3, value=buyer_name, font=_fnt(CALIBRI, 10, bold=True))
        R += 1
        if buyer_address:
            _sc(ws, R, 3, value=buyer_address, font=_fnt(CALIBRI, 9))
            R += 1

        # Row: 空行
        _rh(ws, R, 8)
        R += 1

        # ── 列头区 ──────────────────────────────────────────

        # 判断是吨 / 平方米 / 件
        is_weight_format = model.order.format == "washers_mar"
        is_per_ton = "TON" in model.order.price_unit.upper()
        is_per_sqm = PricingMode.from_price_unit(model.order.price_unit) is PricingMode.PER_SQM

        if is_weight_format or is_per_ton:
            qty_label = "Qty (tons)"
            qty_fmt = "#,##0.00"
        elif is_per_sqm:
            qty_label = "Qty (m²)"
            qty_fmt = "#,##0.00"
        else:
            qty_label = "Qty (pcs)"
            qty_fmt = "#,##0"
        price_label = f"Unit Price\n({model.order.price_unit})"
        price_fmt = "#,##0.00"

        from openpyxl.styles import PatternFill
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = _fnt(CALIBRI, 10, bold=True, color="FFFFFF")
        _sc(ws, R, 1, value="Name of Commodity and Specifications",
            font=hdr_font, align=_aln("left", wrap=True))
        ws.cell(R, 1).fill = hdr_fill
        for ci in range(2, 12):
            ws.cell(R, ci).fill = hdr_fill
        _sc(ws, R, 12, value=qty_label, font=hdr_font, align=_aln("center"))
        ws.cell(R, 12).fill = hdr_fill
        _sc(ws, R, 13, value="Weight\n(kgs)", font=hdr_font, align=_aln("center", wrap=True))
        ws.cell(R, 13).fill = hdr_fill
        _sc(ws, R, 14, value=price_label, font=hdr_font, align=_aln("center", wrap=True))
        ws.cell(R, 14).fill = hdr_fill
        _sc(ws, R, 15, value=f"Total Amount\n({model.order.currency})",
            font=hdr_font, align=_aln("center", wrap=True))
        ws.cell(R, 15).fill = hdr_fill
        _rh(ws, R, 28)
        R += 1

        # 单位行
        _sc(ws, R, 1, value=f"Order No.: {model.order.order_no}",
            font=_fnt(CALIBRI, 11, bold=True))
        _sc(ws, R, 12, border=_brd(left="thin"))
        _sc(ws, R, 13, border=_brd(left="thin"))
        _sc(ws, R, 14, border=_brd(left="thin", right="thin"))
        _sc(ws, R, 15, value=f"{model.order.currency}",
            font=_fnt(CALIBRI, 10), align=_aln("center"))
        R += 1

        # ── 数据区 ──────────────────────────────────────────
        DATA_START = R
        prev_group = None
        col_L = get_column_letter(12)
        col_N = get_column_letter(14)

        for item in items:
            # 品类标题行
            group = item.group_key or item.standard or ""
            if group and group != prev_group:
                _sc(ws, R, 1, value=group,
                    font=_fnt(CALIBRI, 11, bold=True), align=_aln("left"))
                _sc(ws, R, 12, border=_brd(left="thin", right="thin"))
                _sc(ws, R, 13, border=_brd(left="thin", right="thin"))
                _sc(ws, R, 14, border=_brd(left="thin", right="thin"))
                _sc(ws, R, 15, border=_brd(left="thin"))
                _rh(ws, R, 14)
                R += 1
                prev_group = group

            # 数据行
            _sc(ws, R, 1, value=item.description,
                font=_fnt(size=10), align=_aln("left"))

            _sc(ws, R, 12, value=item.quantity,
                font=_fnt(CALIBRI, 11), align=_aln("right"),
                border=_brd(left="thin", right="thin"),
                num_fmt=qty_fmt)

            # Weight 列
            weight = item.weight_kg or (
                item.kg_mpcs if item.kg_mpcs else None
            )
            _sc(ws, R, 13, value=weight,
                font=_fnt(CALIBRI, 11), align=_aln("right"),
                border=_brd(left="thin", right="thin"),
                num_fmt="#,##0.00")

            # Unit Price（从 model 读取，可能为 None）
            _sc(ws, R, 14, value=item.unit_price,
                font=_fnt(CALIBRI, 11), align=_aln("right"),
                border=_brd(left="thin", right="thin"),
                num_fmt=price_fmt)

            # Amount: 统一金额规则（per-mille / per-ton / flat）见 models/amounts.py
            # 列：L=quantity, M=weight, N=unit_price
            col_M = get_column_letter(13)
            formula = amount_formula(
                model.order.price_unit,
                qty_cell=f"{col_L}{R}",
                weight_cell=f"{col_M}{R}",
                price_cell=f"{col_N}{R}",
            )
            _sc(ws, R, 15, value=formula, formula=True,
                font=_fnt(CALIBRI, 11), align=_aln("right"),
                border=_brd(left="thin"),
                num_fmt="#,##0.00")

            _rh(ws, R, 13)
            R += 1

        DATA_END = R - 1

        # ── TOTAL 行 ────────────────────────────────────────
        tb = _brd(top="thin", bottom="thin")
        tblr = _brd(top="thin", bottom="thin", left="thin", right="thin")
        tbl = _brd(top="thin", bottom="thin", left="thin")

        _sc(ws, R, 1, value="TOTAL:",
            font=_fnt(CALIBRI, 10, bold=True), border=tb)
        _sc(ws, R, 12, border=tblr)
        _sc(ws, R, 13, value=f"=SUM(M{DATA_START}:M{DATA_END})", formula=True,
            font=_fnt(CALIBRI, 10, bold=True), align=_aln("right"),
            border=tblr, num_fmt="#,##0.00")
        _sc(ws, R, 14, value=model.order.currency,
            font=_fnt(CALIBRI, 10, bold=True), align=_aln("right"),
            border=tblr)
        _sc(ws, R, 15, value=f"=SUM(O{DATA_START}:O{DATA_END})", formula=True,
            font=_fnt(CALIBRI, 10, bold=True), align=_aln("right"),
            border=tbl, num_fmt="#,##0.00")
        _rh(ws, R, 15)
        R += 1

        # ── 页脚条款 ────────────────────────────────────────
        R += 1

        terms = self.terms
        terms_items = [
            ("Payment Terms:", terms.get("payment", "")),
            ("Delivery:", terms.get("delivery", "")),
            ("Lead Time:", terms.get("lead_time", "")),
            ("Validity:", terms.get("validity", "")),
            ("Packing:", terms.get("packing", "")),
            ("Quality:", terms.get("quality", "")),
        ]
        if terms.get("exchange_rate"):
            terms_items.append(("Exchange Rate:", terms["exchange_rate"]))

        for lbl, val in terms_items:
            _mc(ws, R, 1, R, 4, value=lbl,
                font=_fnt(CALIBRI, 9, bold=True), align=_aln("left", "top"))
            _mc(ws, R, 5, R, 15, value=val,
                font=_fnt(CALIBRI, 9), align=_aln("left", "top", wrap=True))
            _rh(ws, R, 16)
            R += 1

        # 港口
        R += 1
        port_loading = model.derived.port_of_loading or "QINGDAO, CHINA"
        _sc(ws, R, 1, value="Port of Loading:", font=_fnt(CALIBRI, 9, bold=True))
        _sc(ws, R, 5, value=port_loading, font=_fnt(CALIBRI, 9))
        R += 1

        # ── 签字区 ──────────────────────────────────────────
        R += 2
        _sc(ws, R, 1, value="The Buyers:", font=_fnt(CALIBRI, 10))
        _sc(ws, R, 5, value=buyer_name, font=_fnt(CALIBRI, 10, bold=True))
        _sc(ws, R, 12, value="The Sellers:", font=_fnt(CALIBRI, 10))
        _mc(ws, R, 14, R, 15, value=seller_name_en,
            font=_fnt(CALIBRI, 10, bold=True))
        R += 1

        # 买方地址
        for line in buyer_address_lines[:4]:
            _sc(ws, R, 5, value=line, font=_fnt(CALIBRI, 9))
            R += 1

        R += 1
        _mc(ws, R, 1, R, 15,
            value="Authorized Signature: _______________________",
            font=_fnt(CALIBRI, 10), align=_aln("center", "bottom"))

        wb.save(output_path)

        return {
            "items": len(items),
            "total_qty": model.derived.total_qty,
            "output_path": output_path,
            "pi_number": model.order.pi_number,
        }
