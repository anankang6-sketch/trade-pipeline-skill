"""
writers/ci_writer.py — CI 商业发票生成器

从 OrderModel 读取数据，生成 CI Excel。
继承 BaseWriter 统一接口。

与 PI 的区别：
  - 标题: 商业发票 / COMMERCIAL INVOICE
  - 增加: 唛头列、装运地/目的地、信用证号、合约号
  - 增加: 页脚大写金额(SAY)、净重/毛重
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from .base_writer import BaseWriter, sc as _sc, mc as _mc, brd as _brd
from ..models.amounts import amount_formula, compute_amount, PricingMode

CALIBRI = "Calibri"


def _fnt(size=10, bold=False, color="000000", name=CALIBRI):
    return Font(name=name, size=size, bold=bold, color=color)


def _aln(h="left", v=None, wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── 大写金额 ──────────────────────────────────────────────

_ONES = ["", "ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN",
         "EIGHT", "NINE", "TEN", "ELEVEN", "TWELVE", "THIRTEEN",
         "FOURTEEN", "FIFTEEN", "SIXTEEN", "SEVENTEEN", "EIGHTEEN", "NINETEEN"]
_TENS = ["", "", "TWENTY", "THIRTY", "FORTY", "FIFTY",
         "SIXTY", "SEVENTY", "EIGHTY", "NINETY"]


# 量级词从大到小排列，_num_to_words 依次剥离，保证每个量级词只出现一次且合法。
_SCALES = [
    (1_000_000_000_000, "TRILLION"),
    (1_000_000_000, "BILLION"),
    (1_000_000, "MILLION"),
    (1000, "THOUSAND"),
]


def _num_to_words(n: int) -> str:
    if n == 0:
        return "ZERO"
    if n < 0:
        return "MINUS " + _num_to_words(-n)
    parts = []
    for scale, name in _SCALES:
        if n >= scale:
            parts.append(_num_to_words(n // scale) + " " + name)
            n %= scale
    if n >= 100:
        parts.append(_ONES[n // 100] + " HUNDRED")
        n %= 100
    if n >= 20:
        word = _TENS[n // 10]
        if n % 10:
            word += "-" + _ONES[n % 10]
        parts.append(word)
    elif n > 0:
        parts.append(_ONES[n])
    return " ".join(parts)


def amount_to_words(amount: float, currency: str = "USD") -> str:
    """Render a monetary amount as an English words string.

    整数分运算杜绝浮点/边界失真（T2）：
      - total_cents = round(|amount| * 100) 后用 divmod 拆整数位与分位，
        cents 恒在 0..99（0.995 会正确进位到整数位，不再出现 "ONE HUNDRED CENTS"）。
      - 负数统一前缀 MINUS（放在币种之后，如 "USD MINUS FIVE ..."），
        分位由绝对值取得，不会丢失（-5.5 → 5 元 50 分）。
    """
    negative = amount < 0
    total_cents = int(round(abs(amount) * 100))
    integer_part, cents = divmod(total_cents, 100)

    words = f"{currency} "
    if negative and total_cents > 0:
        words += "MINUS "
    words += _num_to_words(integer_part)
    if cents > 0:
        words += " AND " + _num_to_words(cents) + " CENTS"
    words += " ONLY"
    return words


class CIWriter(BaseWriter):
    """CI 商业发票 Writer"""

    def write(self, output_path: str, **kwargs) -> dict:
        model = self.model
        items = model.items
        currency = model.order.currency
        price_unit = model.order.price_unit

        wb = Workbook()
        ws = wb.active
        ws.title = model.order.ci_number or "CI"

        # 列宽 A-R
        col_widths = {
            'A': 7.0, 'B': 11.0, 'C': 4.5, 'D': 3.5, 'E': 3.9, 'F': 3.0,
            'G': 8.43, 'H': 1.9, 'I': 5.5, 'J': 4.9, 'K': 6.9, 'L': 3.9,
            'M': 5.3, 'N': 13.0, 'O': 13.0, 'P': 20.1, 'Q': 33.3, 'R': 12.3,
        }
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w

        # 页面设置
        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = 'landscape'
        ws.page_margins.left = 0.35
        ws.page_margins.right = 0.35
        ws.page_margins.top = 0.47
        ws.page_margins.bottom = 0.51

        seller_name_cn = self.seller.get("name_cn", "")
        seller_name_en = self.seller.get("name_en", "")
        seller_addr = self.seller.get("address", "")
        buyer_name = self.buyer.get("name_en", "") or self.buyer.get("name_ru", "")
        buyer_inn = self.buyer.get("inn", "")
        buyer_address_lines = self.buyer.get("address_lines", [])
        port_loading = model.derived.port_of_loading or "QINGDAO,CHINA"
        port_dest = model.derived.port_of_destination or ""
        ci_date = model.order.date

        R = 1

        # ── Row 1: 中文公司名 ──
        if seller_name_cn:
            _mc(ws, R, 1, R, 17, value=seller_name_cn,
                font=_fnt(24, bold=True, name="宋体"), align=_aln("center", "center"))
            ws.row_dimensions[R].height = 30
            R += 1

        # ── Row 2: 英文公司名 ──
        _mc(ws, R, 1, R, 17, value=seller_name_en,
            font=_fnt(18, bold=True), align=_aln("center", "center"))
        ws.row_dimensions[R].height = 22.5
        R += 1

        # ── Row 3: 地址 ──
        if seller_addr:
            _mc(ws, R, 1, R, 17, value=seller_addr,
                font=_fnt(14), align=_aln("center", "center", wrap=True))
            ws.row_dimensions[R].height = 49.5
            R += 1

        # ── 空行 ──
        ws.row_dimensions[R].height = 12
        R += 1

        # ── 商业发票 ──
        _mc(ws, R, 1, R, 17, value="商业发票",
            font=_fnt(24, bold=True, name="宋体"), align=_aln("center"))
        ws.row_dimensions[R].height = 35.5
        R += 1

        # ── COMMERCIAL INVOICE ──
        _mc(ws, R, 1, R, 17, value="COMMERCIAL INVOICE",
            font=_fnt(20), align=_aln("center"))
        ws.row_dimensions[R].height = 30
        R += 1

        # ── TO + 号码 ──
        _sc(ws, R, 1, value="TO：", font=_fnt(10))
        _sc(ws, R, 2, value=buyer_name, font=_fnt(10))
        _sc(ws, R, 16, value="号码NO.", font=_fnt(11), align=_aln("right"))
        _sc(ws, R, 17, value=model.order.ci_number, font=_fnt(11), align=_aln("right"))
        R += 1

        # ── INN + 日期 ──
        if buyer_inn:
            _sc(ws, R, 2, value=f"INN: {buyer_inn}", font=_fnt(10))
        _sc(ws, R, 16, value="日期DATE", font=_fnt(11), align=_aln("right"))
        _sc(ws, R, 17, value=ci_date, font=_fnt(11), align=_aln("right"))
        R += 1

        # ── 买家地址 + 信用证号 ──
        for i, line in enumerate(buyer_address_lines[:3]):
            _sc(ws, R, 2, value=line, font=_fnt(10))
            if i == 0:
                _sc(ws, R, 16, value="信用证号L/C NO.", font=_fnt(11), align=_aln("right"))
                _sc(ws, R, 17, value=model.order.lc_number or "", font=_fnt(11), align=_aln("right"))
            R += 1

        # ── 合约号 ──
        _sc(ws, R, 16, value="合约号：", font=_fnt(12), align=_aln("right"))
        _sc(ws, R, 17, value="SALES CONTRACT NO.", font=_fnt(11), align=_aln("right"))
        R += 1
        _sc(ws, R, 17, value=model.order.pi_number, font=_fnt(12), align=_aln("right"))
        R += 1

        # ── 装运地/目的地 ──
        _sc(ws, R, 1, value="装运地FROM", font=_fnt(12), border=_brd(bottom="thin"))
        _sc(ws, R, 4, value=port_loading, font=_fnt(11), border=_brd(bottom="thin"))
        _sc(ws, R, 14, value="目的地TO", font=_fnt(11), border=_brd(bottom="thin"))
        _sc(ws, R, 16, value=port_dest, font=_fnt(11), border=_brd(bottom="thin"))
        R += 1

        # ── 分隔 ──
        ws.row_dimensions[R].height = 3.6
        R += 1

        # ── 列头（深蓝底白字，中英文合并为一行） ──
        from openpyxl.styles import PatternFill
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = _fnt(10, bold=True, color="FFFFFF")
        _mc(ws, R, 1, R, 2, value="MARKS & NOS\n唛头", font=hdr_font, align=_aln("center", wrap=True))
        for ci in range(1, 3):
            ws.cell(R, ci).fill = hdr_fill
        _mc(ws, R, 3, R, 13, value="DESCRIPTIONS\n货物名称", font=hdr_font, align=_aln("center", wrap=True))
        for ci in range(3, 14):
            ws.cell(R, ci).fill = hdr_fill
        _sc(ws, R, 14, value="Quantity\n数量", font=hdr_font, align=_aln("center", wrap=True))
        ws.cell(R, 14).fill = hdr_fill
        _sc(ws, R, 15, value="Weight\n重量", font=hdr_font, align=_aln("center", wrap=True))
        ws.cell(R, 15).fill = hdr_fill
        _sc(ws, R, 16, value="Unit Price\n单价", font=hdr_font, align=_aln("center", wrap=True))
        ws.cell(R, 16).fill = hdr_fill
        _sc(ws, R, 17, value="AMOUNT\n金额", font=hdr_font, align=_aln("center", wrap=True))
        ws.cell(R, 17).fill = hdr_fill
        ws.row_dimensions[R].height = 30
        R += 1
        R += 1

        # ── 单位行 ──
        is_per_sqm = PricingMode.from_price_unit(price_unit) is PricingMode.PER_SQM
        unit_label = "tons" if model.order.format == "washers_mar" else ("m²" if is_per_sqm else "pcs")
        _sc(ws, R, 3, value=f"Order No.:{model.order.pi_number}",
            font=_fnt(11, bold=True))
        _sc(ws, R, 14, value=unit_label, font=_fnt(11), align=_aln("right"),
            border=_brd(left="thin"))
        _sc(ws, R, 15, value="kgs", font=_fnt(11), align=_aln("right"),
            border=_brd(left="thin"))
        _sc(ws, R, 16, value=price_unit, font=_fnt(11), align=_aln("right"),
            border=_brd(left="thin", right="thin"))
        _sc(ws, R, 17, value=f"{currency}/ FOB PRICE", font=_fnt(11), align=_aln("center"),
            border=_brd(left="thin"))
        R += 1

        # ── 数据区 ──
        DATA_START = R
        prev_group = None
        total_amount = 0.0

        for item in items:
            # 品类标题行
            group = item.group_key or item.standard or ""
            if group and group != prev_group:
                _sc(ws, R, 3, value=group, font=_fnt(10, bold=True), align=_aln("left"))
                for c in [14, 15, 16]:
                    ws.cell(R, c).border = _brd(left="thin")
                ws.row_dimensions[R].height = 14.1
                R += 1
                prev_group = group

            # 数据行
            _sc(ws, R, 3, value=item.description, font=_fnt(10), align=_aln("left"))

            # N: qty
            _sc(ws, R, 14, value=item.quantity, font=_fnt(10), align=_aln("right"),
                border=_brd(left="thin"),
                num_fmt='#,##0.00' if is_per_sqm else '#,##0')

            # O: weight
            weight = item.weight_kg
            _sc(ws, R, 15, value=weight, font=_fnt(10), align=_aln("right"),
                border=_brd(left="thin"), num_fmt='#,##0.00')

            # P: unit price
            _sc(ws, R, 16, value=item.unit_price, font=_fnt(10), align=_aln("right"),
                border=_brd(left="thin", right="thin"), num_fmt='#,##0.00')

            # Q: amount formula — 统一金额规则见 models/amounts.py
            # 列：N=qty, O=weight, P=unit_price
            col_P = get_column_letter(16)
            col_N = get_column_letter(14)
            col_O = get_column_letter(15)
            formula = amount_formula(
                price_unit,
                qty_cell=f"{col_N}{R}",
                weight_cell=f"{col_O}{R}",
                price_cell=f"{col_P}{R}",
            )
            _sc(ws, R, 17, value=formula, formula=True,
                font=_fnt(10), align=_aln("right"), num_fmt='#,##0.00')

            # 累计金额（与公式同源，杜绝显示值与大写金额不一致）
            total_amount += compute_amount(
                price_unit,
                quantity=item.quantity,
                weight_kg=item.weight_kg,
                unit_price=item.unit_price,
            )

            ws.row_dimensions[R].height = 14.1
            R += 1

        DATA_END = R - 1

        # ── 空行 ──
        ws.row_dimensions[R].height = 14.1
        R += 1

        # ── TOTAL 行 ──
        _sc(ws, R, 3, value="TOTAL:", font=_fnt(12), align=_aln("left"))
        _sc(ws, R, 14, value=f"=SUM(N{DATA_START}:N{DATA_END})", formula=True,
            font=_fnt(12), align=_aln("right"),
            border=_brd(left="thin", top="thin", bottom="thin"),
            num_fmt='#,##0.00' if is_per_sqm else '#,##0')
        _sc(ws, R, 15, value=f"=SUM(O{DATA_START}:O{DATA_END})", formula=True,
            font=_fnt(12), align=_aln("right"),
            border=_brd(left="thin", top="thin", bottom="thin"), num_fmt='#,##0.00')
        _sc(ws, R, 16, value=currency, font=_fnt(12), align=_aln("right"),
            border=_brd(left="thin", right="thin", top="thin", bottom="thin"))
        _sc(ws, R, 17, value=f"=SUM(Q{DATA_START}:Q{DATA_END})", formula=True,
            font=_fnt(12), align=_aln("right"),
            border=_brd(left="thin", top="thin", bottom="thin"), num_fmt='#,##0.00')
        ws.row_dimensions[R].height = 18.6
        R += 1

        # ── SAY 大写金额 ──
        has_unpriced = any(item.unit_price is None for item in items)
        if total_amount == 0 and has_unpriced:
            say_text = "SAY: (PRICES TO BE CONFIRMED)"
        else:
            say_text = f"SAY: {amount_to_words(total_amount, currency)}."
        _sc(ws, R, 1, value=say_text, font=_fnt(12))
        R += 1

        # ── 托盘信息（占位，PL 完成后可回填）──
        pallet_count = model.derived.pallet_count
        if pallet_count:
            _sc(ws, R, 1, value=f"PACKED IN {pallet_count} PALLETS ONLY.", font=_fnt(12))
        else:
            _sc(ws, R, 1, value="PACKED IN [    ] PALLETS ONLY.", font=_fnt(12))
        R += 1

        # ── 净重/毛重 ──
        nw = sum(i.weight_kg or 0 for i in items)
        gw = model.derived.total_gross_weight or round(nw * 1.036, 2)
        _sc(ws, R, 1, value=f"N.W.:{nw:,.2f}KGS  G.W.:{gw:,.2f}KGS", font=_fnt(12))
        R += 1

        wb.save(output_path)

        return {
            "items": len(items),
            "ci_number": model.order.ci_number,
            "total_amount": round(total_amount, 2),
            "total_net_weight": round(nw, 2),
            # T1: 暴露 CI 页脚的毛重（读 derived 真值或走兜底），供跨单校验比对
            "total_gross_weight": round(gw, 2),
            "output_path": output_path,
        }
