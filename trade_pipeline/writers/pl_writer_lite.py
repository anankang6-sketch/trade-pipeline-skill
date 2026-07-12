"""
writers/pl_writer_lite.py — Simplified PL generator for public demo

Generates a basic Packing List from OrderModel without requiring the
private pl-gen engine. Uses simplified packing rules:
  - Fixed 25 kg per carton
  - 36 cartons per pallet
  - Net weight from item data, gross weight = net + pallet self-weight

Production environments use the separate pl-gen engine with
customer-specific templates and packing rules.
"""
import math
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .base_writer import BaseWriter, sc as _sc, mc as _mc, brd as _brd

CALIBRI = "Calibri"
DEFAULT_KG_PER_CARTON = 25.0
DEFAULT_CARTONS_PER_PALLET = 36
DEFAULT_PALLET_SELF_WEIGHT_KG = 28.0
DEFAULT_MEASUREMENT_PER_PALLET_M3 = 0.528
HEADER_COLOR = "1F3864"
BORDER_COLOR = "E0E0E0"


def _fnt(size=10, bold=False, color="000000"):
    return Font(name=CALIBRI, size=size, bold=bold, color=color)


def _aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


@dataclass
class PackingLine:
    description: str
    group_key: str
    pcs: float                # 面积（㎡）
    net_weight_kg: float
    pallets: int
    kg_per_pallet: float
    pallet_type: str = "wooden"
    gross_weight_kg: float = 0.0


def _compute_packing(
    items,
    pallet_presets: dict,
    default_pallet: str = "wooden",
) -> tuple[list[PackingLine], dict]:
    """按行分配托盘类型，逐托计算装箱（落石防护网版）。

    每个 item 的 area = quantity（㎡），pallet_type 决定用哪种托盘预设
    （wooden / metal），预设含 capacity_m2 / self_weight_kg / measurement_m3。
    重量仅用于装箱，不参与计价。
    """
    lines = []
    total_area = 0.0
    total_net = 0.0
    total_gross = 0.0
    total_measurement = 0.0
    total_pallets = 0

    for item in items:
        # 总净重优先级：weight_kg → weight_kg_per_piece * qty → kg_mpcs * qty/1000
        # 用 `is not None` 而非 falsy 检查，0.0 是合法重量（赠品/试样）
        nw = item.weight_kg
        if nw is None:
            wpp = getattr(item, "weight_kg_per_piece", None)
            if wpp is not None:
                nw = wpp * item.quantity
            elif item.kg_mpcs is not None:
                nw = item.kg_mpcs * item.quantity / 1000
            else:
                nw = 0

        area = item.quantity or 0.0
        ptype = getattr(item, "pallet_type", None) or default_pallet
        preset = pallet_presets.get(ptype) or pallet_presets.get(default_pallet) or {}
        cap = float(preset.get("capacity_m2", 200.0))
        self_w = float(preset.get("self_weight_kg", 28.0))
        meas = float(preset.get("measurement_m3", 0.528))

        # 按面积算托盘数（每托盘可铺 capacity_m2 ㎡）
        num_pallets = max(1, math.ceil(area / cap)) if area > 0 else 1
        kg_per_pallet = round(nw / num_pallets, 2) if num_pallets > 0 else 0.0
        gross = nw + num_pallets * self_w
        measurement = num_pallets * meas

        lines.append(PackingLine(
            description=item.description,
            group_key=item.group_key or "",
            pcs=area,
            net_weight_kg=round(nw, 2),
            pallets=num_pallets,
            kg_per_pallet=kg_per_pallet,
            pallet_type=ptype,
            gross_weight_kg=round(gross, 2),
        ))
        total_area += area
        total_net += nw
        total_pallets += num_pallets
        total_gross += gross
        total_measurement += measurement

    summary = {
        "total_pcs": total_area,
        "total_net_weight": round(total_net, 2),
        "total_gross_weight": round(total_gross, 2),
        "total_pallets": total_pallets,
        "total_measurement_m3": round(total_measurement, 2),
    }
    return lines, summary


class PackingInfoMissingError(Exception):
    """Raised when items lack weight info needed for PL generation."""

    def __init__(self, missing_items: list):
        self.missing_items = missing_items
        names = ", ".join(f"#{it.no} {it.description[:40]}" for it in missing_items[:3])
        more = f" 等 {len(missing_items)} 项" if len(missing_items) > 3 else ""
        super().__init__(
            f"PL 生成缺少重量信息: {names}{more}。"
            "请补充每个规格的单件重量 (weight_kg) 或每千件重量 (kg_mpcs)，"
            "或在 GUI 中通过装箱向导填写。"
        )


def _check_weight_completeness(items) -> list:
    """Return list of items missing weight data needed for PL.

    v1.1.0: 也认 weight_kg_per_piece（Gateway 收集）和 pcs_per_carton（按件包装）。
    pcs_per_carton 单独存在时仍需 weight 信息，但 weight_kg_per_piece 是新主路径。

    Bug fix (C4): 用 `is not None` 区分 "用户没提供重量数据" 和 "用户提供了 0kg
    （如赠品/试样/包装内附件）"。前者应该报 missing，后者是合法值不应该报。
    """
    missing = []
    for item in items:
        # 有任何一个 weight 字段非 None 都视为已提供（即使值是 0）
        has_weight_source = (
            item.weight_kg is not None
            or getattr(item, "weight_kg_per_piece", None) is not None
            or item.kg_mpcs is not None
        )
        if not has_weight_source:
            missing.append(item)
    return missing


class PLWriterLite(BaseWriter):
    """Simplified PL Writer for public demo.

    v1.1.0 新增 kwargs:
      - allow_missing_weight=True  : 跳过 safety net（demo / 测试用）
      - packing_review=PackingReview : 用 review.pallet 覆盖默认 pallet 配置
    """

    def write(self, output_path: str, **kwargs) -> dict:
        model = self.model
        items = model.items

        # ── Safety net: detect missing weight info before generating empty PL ──
        # Skipped if caller explicitly accepts zero-weight output (e.g. demo with all zeros).
        if not kwargs.get("allow_missing_weight", False):
            missing = _check_weight_completeness(items)
            if missing:
                raise PackingInfoMissingError(missing)

        # ── Packing review override (v1.1.0) ──
        # 如果上层传入 PackingReview（已 apply 到 model），用它的 pallet 配置覆盖
        review = kwargs.get("packing_review")

        packing_cfg = (self.config or {}).get("packing", {})
        # 逐行托盘预设：config.packing.pallets.{wooden,metal}
        pallet_presets = packing_cfg.get("pallets", {})
        if not pallet_presets:
            # 兼容：无预设时退回模块默认木托盘
            pallet_presets = {
                "wooden": {
                    "capacity_m2": 200.0,
                    "self_weight_kg": DEFAULT_PALLET_SELF_WEIGHT_KG,
                    "measurement_m3": DEFAULT_MEASUREMENT_PER_PALLET_M3,
                }
            }
        packing_lines, summary = _compute_packing(items, pallet_presets)

        wb = Workbook()
        ws = wb.active
        ws.title = "PACKING LIST"

        col_widths = {"A": 6, "B": 40, "C": 10, "D": 12, "E": 12, "F": 14, "G": 14}
        for letter, w in col_widths.items():
            ws.column_dimensions[letter].width = w

        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = "landscape"

        seller_name_cn = self.seller.get("name_cn", "")
        seller_name_en = self.seller.get("name_en", "")
        buyer_name = self.buyer.get("name_en", "") or self.buyer.get("name_ru", "")
        port_loading = model.derived.port_of_loading or "QINGDAO, CHINA"

        R = 1

        # Header
        if seller_name_cn:
            _mc(ws, R, 1, R, 7, value=seller_name_cn,
                font=Font(name="宋体", size=18, bold=True), align=_aln("center"))
            R += 1

        _mc(ws, R, 1, R, 7, value=seller_name_en,
            font=_fnt(14, bold=True), align=_aln("center"))
        R += 1

        R += 1
        _mc(ws, R, 1, R, 7, value="PACKING LIST",
            font=_fnt(16, bold=True), align=_aln("center"))
        R += 1

        R += 1
        _sc(ws, R, 1, value="TO:", font=_fnt(10))
        _sc(ws, R, 2, value=buyer_name, font=_fnt(10, bold=True))
        _sc(ws, R, 6, value="PI No.:", font=_fnt(10))
        _sc(ws, R, 7, value=model.order.pi_number, font=_fnt(10, bold=True))
        R += 1

        _sc(ws, R, 6, value="Date:", font=_fnt(10))
        _sc(ws, R, 7, value=model.order.date, font=_fnt(10))
        R += 1

        _sc(ws, R, 1, value=f"FROM: {port_loading}", font=_fnt(10))
        R += 1

        R += 1

        # Column headers
        headers = ["No.", "Description", "Pallets", "KGS/PALLET", "Qty (m²)", "Net Weight", "Gross Weight"]
        header_fill = PatternFill("solid", fgColor="1F3864")
        for i, h in enumerate(headers, 1):
            c = _sc(ws, R, i, value=h,
                    font=Font(name=CALIBRI, size=10, bold=True, color="FFFFFF"),
                    align=_aln("center", wrap=True))
            c.fill = header_fill
        R += 1

        # Data rows
        prev_group = None
        thin_bottom = _brd(bottom="thin")

        for idx, line in enumerate(packing_lines, 1):
            if line.group_key and line.group_key != prev_group:
                _sc(ws, R, 2, value=line.group_key, font=_fnt(10, bold=True))
                R += 1
                prev_group = line.group_key

            # 描述后缀托盘类型，便于核对（WOODEN / METAL）
            desc_cell = f"{line.description}  [{line.pallet_type.upper()}]"
            _sc(ws, R, 1, value=idx, font=_fnt(10), align=_aln("center"), border=thin_bottom)
            _sc(ws, R, 2, value=desc_cell, font=_fnt(10), border=thin_bottom)
            _sc(ws, R, 3, value=line.pallets, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0")
            _sc(ws, R, 4, value=line.kg_per_pallet, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0.00")
            _sc(ws, R, 5, value=line.pcs, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0.00")
            _sc(ws, R, 6, value=line.net_weight_kg, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0.00")
            _sc(ws, R, 7, value=line.gross_weight_kg, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0.00")
            R += 1

        # Total row
        total_border = _brd(top="thin", bottom="thin")
        _sc(ws, R, 1, value="", border=total_border)
        _sc(ws, R, 2, value="TOTAL:", font=_fnt(11, bold=True), border=total_border)
        _sc(ws, R, 3, value=summary["total_pallets"], font=_fnt(11, bold=True),
            align=_aln("right"), border=total_border, num_fmt="#,##0")
        _sc(ws, R, 4, value="", border=total_border)
        _sc(ws, R, 5, value=summary["total_pcs"], font=_fnt(11, bold=True),
            align=_aln("right"), border=total_border, num_fmt="#,##0.00")
        _sc(ws, R, 6, value=summary["total_net_weight"], font=_fnt(11, bold=True),
            align=_aln("right"), border=total_border, num_fmt="#,##0.00")
        _sc(ws, R, 7, value=summary["total_gross_weight"], font=_fnt(11, bold=True),
            align=_aln("right"), border=total_border, num_fmt="#,##0.00")
        R += 1

        # Footer
        R += 1
        _sc(ws, R, 1, value=f"PACKED IN {summary['total_pallets']} PALLETS ONLY.",
            font=_fnt(10))
        R += 1
        _sc(ws, R, 1, value=f"TOTAL MEASUREMENT: {summary['total_measurement_m3']:.2f} m³",
            font=_fnt(10))
        R += 1
        _sc(ws, R, 1, value="PACKING: ON WOODEN / METAL PALLETS, STRETCH FILM.",
            font=_fnt(10))
        R += 1
        R += 1
        _sc(ws, R, 1, value=f"N.W.: {summary['total_net_weight']:,.2f} KGS   "
                             f"G.W.: {summary['total_gross_weight']:,.2f} KGS",
            font=_fnt(10, bold=True))

        wb.save(output_path)

        # Write back to model
        model.derived.pallet_count = summary["total_pallets"]
        model.derived.total_net_weight = summary["total_net_weight"]
        model.derived.total_gross_weight = summary["total_gross_weight"]
        model.derived.total_measurement_m3 = summary["total_measurement_m3"]

        return {
            "success": True,
            "pl_path": output_path,
            "items": len(packing_lines),
            "total_pallets": summary["total_pallets"],
            "total_net_weight": summary["total_net_weight"],
            "total_gross_weight": summary["total_gross_weight"],
        }
