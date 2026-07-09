"""Tests for UUID-anchored price write-back.

Core invariant: price write-back must work correctly even when
users insert rows, add group headers, or reorder items in the quotation.
"""
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from trade_pipeline.models.order_model import OrderModel, OrderRefs, OrderInfo, OrderItem, DerivedData, OrderMeta
from trade_pipeline.writers.quote_writer import write_quotation_with_uuid
from trade_pipeline.pipeline.price_updater import update_prices


def _make_items():
    return [
        OrderItem(no=1, item_uuid="aaa111", part_no=None, standard="DIN 933",
                  description_raw="BOLT M8", description="HEX BOLT M8x25",
                  quantity=1000, unit="pcs"),
        OrderItem(no=2, item_uuid="bbb222", part_no=None, standard="DIN 934",
                  description_raw="NUT M8", description="HEX NUT M8",
                  quantity=2000, unit="pcs"),
        OrderItem(no=3, item_uuid="ccc333", part_no=None, standard="DIN 125",
                  description_raw="WASHER M8", description="FLAT WASHER M8",
                  quantity=5000, unit="pcs"),
    ]


def _make_model(items):
    return OrderModel(
        refs=OrderRefs(seller_id="test", buyer_id="test", terms_id="test"),
        order=OrderInfo(order_no="T1", format="standard", quote_no="QT-T1",
                        pi_number="PI-T1", ci_number="CI-T1", date="19 May 2026",
                        currency="USD", price_unit="USD/PC"),
        items=items,
        derived=DerivedData(total_items=len(items), total_qty=sum(i.quantity for i in items)),
        meta=OrderMeta(),
    )


def test_normal_price_writeback():
    items = _make_items()
    model = _make_model(items)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_quotation_with_uuid(items, path, "T1")

        wb = load_workbook(path)
        ws = wb.active
        # Find data rows by scanning for uuid column
        uuid_col = None
        for c in range(1, ws.max_column + 1):
            for r in range(1, 20):
                if ws.cell(r, c).value == "__item_uuid__":
                    uuid_col = c
                    break
            if uuid_col:
                break

        # Fill prices: find rows with uuids and fill price column (col 6)
        for r in range(1, ws.max_row + 1):
            uid = ws.cell(r, uuid_col).value
            if uid == "aaa111":
                ws.cell(r, 6, 10.50)
            elif uid == "bbb222":
                ws.cell(r, 6, 5.00)
            elif uid == "ccc333":
                ws.cell(r, 6, 2.25)
        wb.save(path)
        wb.close()

        result = update_prices(model, path)
        assert result["updated"] == 3
        assert not result["errors"]
        assert model.items[0].unit_price == 10.50
        assert model.items[1].unit_price == 5.00
        assert model.items[2].unit_price == 2.25
    finally:
        Path(path).unlink(missing_ok=True)


def test_writeback_after_row_insertion():
    """User inserts 3 blank rows at the top of data area — prices still land correctly."""
    items = _make_items()
    model = _make_model(items)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_quotation_with_uuid(items, path, "T1")

        wb = load_workbook(path)
        ws = wb.active

        uuid_col = None
        for c in range(1, ws.max_column + 1):
            for r in range(1, 20):
                if ws.cell(r, c).value == "__item_uuid__":
                    uuid_col = c
                    break
            if uuid_col:
                break

        # Insert 3 rows after header (before data), shifting data down
        header_row = None
        for r in range(1, 20):
            if ws.cell(r, uuid_col).value == "__item_uuid__":
                header_row = r
                break
        ws.insert_rows(header_row + 1, 3)

        # Now fill prices (data rows shifted down by 3)
        for r in range(1, ws.max_row + 1):
            uid = ws.cell(r, uuid_col).value
            if uid == "aaa111":
                ws.cell(r, 6, 99.99)
            elif uid == "bbb222":
                ws.cell(r, 6, 88.88)
            elif uid == "ccc333":
                ws.cell(r, 6, 77.77)
        wb.save(path)
        wb.close()

        result = update_prices(model, path)
        assert result["updated"] == 3
        assert not result["errors"]
        assert model.items[0].unit_price == 99.99
        assert model.items[1].unit_price == 88.88
        assert model.items[2].unit_price == 77.77
    finally:
        Path(path).unlink(missing_ok=True)


def test_writeback_with_missing_uuid():
    """If a model item's UUID is not found in Excel, it should be reported."""
    items = _make_items()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_quotation_with_uuid(items, path, "T1")

        # Now create a model with a different uuid for item 3
        model = _make_model(_make_items())
        model.items[2].item_uuid = "not_in_excel"

        wb = load_workbook(path)
        ws = wb.active
        uuid_col = None
        for c in range(1, ws.max_column + 1):
            for r in range(1, 20):
                if ws.cell(r, c).value == "__item_uuid__":
                    uuid_col = c
                    break
            if uuid_col:
                break

        for r in range(1, ws.max_row + 1):
            uid = ws.cell(r, uuid_col).value
            if uid == "aaa111":
                ws.cell(r, 6, 10.0)
            elif uid == "bbb222":
                ws.cell(r, 6, 20.0)
        wb.save(path)
        wb.close()

        result = update_prices(model, path)
        assert result["updated"] == 2
        assert len(result["errors"]) >= 1
    finally:
        Path(path).unlink(missing_ok=True)


def test_missing_price_column_raises_instead_of_fallback():
    """遗漏4 修复的回归防线：找不到单价列必须硬阻断（PriceUpdateError），
    绝不能 fallback 到固定列号从错误列读数当单价回写。"""
    import pytest
    from openpyxl import Workbook
    from trade_pipeline.pipeline.price_updater import PriceUpdateError

    model = _make_model(_make_items())

    wb = Workbook()
    ws = wb.active
    # 有 uuid 列，但表头没有任何含 'Price' 的列
    ws.cell(1, 1, "Item")
    ws.cell(1, 2, "__item_uuid__")
    ws.cell(2, 1, "HEX BOLT M8x25")
    ws.cell(2, 2, "aaa111")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        wb.close()
        with pytest.raises(PriceUpdateError, match="单价列"):
            update_prices(model, path)
    finally:
        Path(path).unlink(missing_ok=True)
